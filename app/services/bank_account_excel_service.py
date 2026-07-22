import openpyxl
from openpyxl.styles import Font, PatternFill

from app.database.connection import get_session
from app.database.models import BankAccount, Member
from app.services.bank_account_service import BankAccountService


HEADERS = [
    "顧客コード", "顧客名", "口座ID", "使用可否", "金融機関コード", "金融機関名",
    "支店コード", "支店名", "預金種目コード", "預金種目", "口座番号", "受取人名カナ",
]
ACCOUNT_TYPE_NAMES = {"1": "普通", "2": "当座", "4": "貯蓄"}


def _code(value, length: int) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text.zfill(length) if text.isdigit() else text


def _enabled(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value or "").strip().lower() in {"1", "true", "yes", "有効", "使用可", "○"}


class BankAccountExcelService:
    def __init__(self, engine):
        self._engine = engine
        self._accounts = BankAccountService(engine)

    def export_excel(self, output_path: str, member_ids: list[int] | None = None) -> int:
        with get_session(self._engine) as session:
            query = session.query(BankAccount, Member).join(Member, BankAccount.member_id == Member.id)
            if member_ids is not None:
                query = query.filter(Member.id.in_(member_ids))
            rows = query.order_by(Member.company_code, BankAccount.id).all()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "振込先口座"
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="3D5A80")
            for account, member in rows:
                ws.append([
                    member.company_code, member.org_name, account.id,
                    "有効" if account.is_enabled else "無効",
                    account.bank_code, account.bank_name, account.branch_code,
                    account.branch_name, account.account_type,
                    ACCOUNT_TYPE_NAMES.get(account.account_type, ""), account.account_number,
                    account.recipient_name_kana,
                ])
            for column in ("E", "G", "I", "K"):
                for cell in ws[column][1:]:
                    cell.number_format = "@"
            widths = [12, 28, 10, 10, 16, 24, 12, 24, 16, 12, 14, 30]
            for index, width in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            wb.save(output_path)
            return len(rows)

    @staticmethod
    def generate_template(output_path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "振込先口座"
        ws.append(HEADERS)
        ws.append([
            1001, "（確認用・入力任意）", "", "有効", "0155", "百五銀行",
            "307", "旭が丘支店", "1", "普通", "0123456", "ｶ)ｻﾝﾌﾟﾙ",
        ])
        ws["A4"] = "入力方法"
        ws["A5"] = "顧客コードは必須です。口座IDが空欄なら追加、既存IDなら更新します。"
        ws["A6"] = "使用可否は「有効」「無効」、預金種目コードは 1:普通 2:当座 4:貯蓄です。"
        for column in ("E", "G", "I", "K"):
            for cell in ws[column]:
                cell.number_format = "@"
        wb.save(output_path)

    def import_excel(self, path: str) -> dict:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook["振込先口座"] if "振込先口座" in workbook.sheetnames else workbook.active
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        missing = [name for name in HEADERS if name not in headers]
        if missing:
            raise ValueError("必要な列がありません：" + "、".join(missing))
        indexes = {name: headers.index(name) for name in HEADERS}
        result = {"added": 0, "updated": 0, "skipped": 0, "errors": []}
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(value not in (None, "") for value in row):
                continue
            try:
                raw_company_code = row[indexes["顧客コード"]]
                company_code = int(raw_company_code) if raw_company_code not in (None, "") else None
                if company_code is None:
                    raise ValueError("顧客コードを入力してください。")
                with get_session(self._engine) as session:
                    member = session.query(Member).filter_by(company_code=company_code).first()
                    if not member:
                        raise ValueError(f"顧客コード{company_code}が見つかりません。")
                    member_id = member.id
                data = {
                    "is_enabled": _enabled(row[indexes["使用可否"]]),
                    "bank_code": _code(row[indexes["金融機関コード"]], 4),
                    "bank_name": str(row[indexes["金融機関名"]] or "").strip(),
                    "branch_code": _code(row[indexes["支店コード"]], 3),
                    "branch_name": str(row[indexes["支店名"]] or "").strip(),
                    "account_type": _code(row[indexes["預金種目コード"]], 1),
                    "account_number": _code(row[indexes["口座番号"]], 7),
                    "recipient_name_kana": str(row[indexes["受取人名カナ"]] or "").strip(),
                }
                raw_id = row[indexes["口座ID"]]
                if raw_id not in (None, ""):
                    account_id = int(raw_id)
                    existing = self._accounts.get(account_id)
                    if not existing or existing.member_id != member_id:
                        raise ValueError("口座IDがこの顧客の口座と一致しません。")
                    self._accounts.update(account_id, data)
                    result["updated"] += 1
                else:
                    self._accounts.create(member_id, data)
                    result["added"] += 1
            except Exception as exc:
                result["skipped"] += 1
                result["errors"].append(f"{row_number}行目：{exc}")
        return result
