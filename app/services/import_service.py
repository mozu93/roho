# app/services/import_service.py
import openpyxl
from datetime import date as _date
from app.database.connection import get_session
from app.database.models import Member, InsuranceEntry
from app.services.member_service import MemberService
from app.utils.kana import to_halfwidth_kana


def _parse_date(val) -> "_date | None":
    """セル値を date に変換。datetime/date はそのまま、文字列は YYYY-MM-DD or YYYY/MM/DD を解析。"""
    if val is None:
        return None
    if isinstance(val, _date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# Excel列インデックス（0始まり）→フィールドマッピング
DEFAULT_COL_MAP = {
    "company_code":      0,   # A  管理No.（上書き時の照合キー）
    "is_member":         1,   # B  ○=会員、空=非会員
    "member_number":     2,   # C
    "org_name":          3,   # D
    "org_kana":          4,   # E
    "dept_title":        5,   # F
    "rep_name":          6,   # G
    "rep_kana":          7,   # H
    "email":             8,   # I
    "tel_area":          9,   # J
    "tel":              10,   # K
    "fax_area":         11,   # L
    "fax":              12,   # M
    "postal_code":      13,   # N
    "address":          14,   # O
    "address2":         15,   # P  建物名等
    "postal_code_mail": 16,   # Q
    "address_mail":     17,   # R
    "address_mail2":    18,   # S  郵送先建物名等
    "mail_org_name":    19,   # T
    "mail_dept_title":  20,   # U
    "mail_person_name": 21,   # V
    "employment_ins_no": 22,  # W
    # 保険番号（X=23始まり、各2列＋フラグ2列）
    "ins_ippan_branch":              23,
    "ins_ippan_number":              24,
    "ins_ippan_tokubetsu":           25,
    "ins_ippan_ikkatsu":             26,
    "ins_kensetsu_koyou_branch":     27,
    "ins_kensetsu_koyou_number":     28,
    "ins_kensetsu_koyou_tokubetsu":  29,
    "ins_kensetsu_koyou_ikkatsu":    30,
    "ins_ringyo_branch":             31,
    "ins_ringyo_number":             32,
    "ins_ringyo_tokubetsu":          33,
    "ins_ringyo_ikkatsu":            34,
    "ins_kensetsu_genba_branch":     35,
    "ins_kensetsu_genba_number":     36,
    "ins_kensetsu_genba_tokubetsu":  37,
    "ins_kensetsu_genba_ikkatsu":    38,
    "ins_kensetsu_jimusho_branch":   39,
    "ins_kensetsu_jimusho_number":   40,
    "ins_kensetsu_jimusho_tokubetsu": 41,
    "ins_kensetsu_jimusho_ikkatsu":  42,
    "note":              43,
    "registered_date":   44,  # AS
    "email_1_address":   45,
    "email_1_label":     46,
    "email_2_address":   47,
    "email_2_label":     48,
    "email_3_address":   49,
    "email_3_label":     50,
}

# 旧テンプレート（「郵送先宛名」1列）の取込互換用。旧宛名は郵送先事業所名へ移す。
LEGACY_DEFAULT_COL_MAP = {
    key: (idx - 2 if idx >= 22 else idx)
    for key, idx in DEFAULT_COL_MAP.items()
    if key not in ("mail_dept_title", "mail_person_name")
}
LEGACY_DEFAULT_COL_MAP["mail_org_name"] = 19

INS_TYPE_KEYS = [
    ("ippan",            "ins_ippan"),
    ("kensetsu_koyou",   "ins_kensetsu_koyou"),
    ("ringyo",           "ins_ringyo"),
    ("kensetsu_genba",   "ins_kensetsu_genba"),
    ("kensetsu_jimusho", "ins_kensetsu_jimusho"),
]
BRANCH_NUMBERS = {
    "ippan": "0", "kensetsu_koyou": "2",
    "ringyo": "4", "kensetsu_genba": "5", "kensetsu_jimusho": "6",
}

EXPORT_HEADERS = [
    "管理No.", "会", "会員No.", "事業所名", "フリガナ", "所属・役職",
    "代表者名", "代表者フリガナ", "メール", "市外局番", "電話番号", "FAX市外局番", "FAX",
    "郵便番号", "住所", "建物名等",
    "郵送先郵便番号", "郵送先住所", "郵送先建物名等",
    "郵送先事業所名", "郵送先所属・役職名", "郵送先氏名", "雇用保険事業所番号",
    "一般枝番", "一般番号", "一般特別加入", "一般一括",
    "建設他雇枝番", "建設他雇番号", "建設他雇特別", "建設他雇一括",
    "林業枝番", "林業番号", "林業特別", "林業一括",
    "建設現場枝番", "建設現場番号", "建設現場特別", "建設現場一括",
    "建設事務所枝番", "建設事務所番号", "建設事務所特別", "建設事務所一括",
    "メモ",
    "登録日",
    "メール1アドレス", "メール1ラベル",
    "メール2アドレス", "メール2ラベル",
    "メール3アドレス", "メール3ラベル",
]


class ImportService:
    def __init__(self, engine):
        self._engine = engine
        self._svc = MemberService(engine)

    def import_excel(
        self,
        path: str,
        col_map: dict | None = None,
        overwrite: bool = False,
        staff_name: str = "インポート",
    ) -> dict:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        if col_map is None:
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            is_legacy = len(headers) > 19 and str(headers[19] or "").strip() == "郵送先宛名"
            col_map = LEGACY_DEFAULT_COL_MAP if is_legacy else DEFAULT_COL_MAP
        result = {"added": 0, "updated": 0, "skipped": 0, "skipped_details": []}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            member_number = str(row[col_map["member_number"]] or "").strip() or None
            org_name = str(row[col_map["org_name"]] or "").strip()
            if not org_name:
                result["skipped"] += 1
                result["skipped_details"].append(f"{row_idx}行目：事業所名が空のためスキップ")
                continue

            def _get(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            ins_entries = []
            for ins_type, prefix in INS_TYPE_KEYS:
                branch = str(_get(f"{prefix}_branch") or "").strip()
                number = str(_get(f"{prefix}_number") or "").strip()
                if branch or number:
                    ins_entries.append({
                        "ins_type": ins_type,
                        "branch_number": branch or BRANCH_NUMBERS[ins_type],
                        "ins_number": number,
                        "is_tokubetsu": bool(_get(f"{prefix}_tokubetsu")),
                        "is_ikkatsu": bool(_get(f"{prefix}_ikkatsu")),
                    })

            raw_is_member = _get("is_member")
            if isinstance(raw_is_member, str):
                is_member = raw_is_member.strip() in ("○", "true", "True", "1")
            elif isinstance(raw_is_member, bool):
                is_member = raw_is_member
            elif raw_is_member is not None:
                is_member = bool(raw_is_member)
            else:
                is_member = False

            # 住所: 1行目と建物名等を \n 結合
            addr1 = str(_get("address") or "").strip()
            addr2 = str(_get("address2") or "").strip()
            addr = "\n".join(x for x in [addr1, addr2] if x)

            addr_mail1 = str(_get("address_mail") or "").strip()
            addr_mail2 = str(_get("address_mail2") or "").strip()
            addr_mail = "\n".join(x for x in [addr_mail1, addr_mail2] if x)

            email_addresses = []
            for i in range(1, 4):
                email_address = str(_get(f"email_{i}_address") or "").strip()
                if email_address:
                    email_addresses.append({
                        "address": email_address,
                        "label": str(_get(f"email_{i}_label") or "").strip(),
                    })
            if not email_addresses:
                legacy_email = str(_get("email") or "").strip()
                if legacy_email:
                    email_addresses.append({"address": legacy_email, "label": ""})

            data = {
                "member_number":     member_number,
                "is_member":         is_member,
                "org_name":          org_name,
                "org_kana":          to_halfwidth_kana(str(_get("org_kana") or "")),
                "dept_title":        str(_get("dept_title") or ""),
                "rep_name":          str(_get("rep_name") or ""),
                "rep_kana":          to_halfwidth_kana(str(_get("rep_kana") or "")),
                "email_addresses":   email_addresses,
                "tel_area":          str(_get("tel_area") or ""),
                "tel":               str(_get("tel") or ""),
                "fax_area":          str(_get("fax_area") or ""),
                "fax":               str(_get("fax") or ""),
                "postal_code":       str(_get("postal_code") or ""),
                "address":           addr,
                "postal_code_mail":  str(_get("postal_code_mail") or ""),
                "address_mail":      addr_mail,
                "mail_org_name":     str(_get("mail_org_name") or ""),
                "mail_dept_title":   str(_get("mail_dept_title") or ""),
                "mail_person_name":  str(_get("mail_person_name") or ""),
                "employment_ins_no": str(_get("employment_ins_no") or ""),
                "note":              str(_get("note") or ""),
                "registered_date":   _parse_date(_get("registered_date")),
                "insurance_entries": ins_entries,
            }

            # 既存チェック（管理No.があればそれで、なければ事業所名で照合）
            raw_cc = _get("company_code")
            company_code = int(raw_cc) if raw_cc else None
            with get_session(self._engine) as session:
                if company_code:
                    existing = session.query(Member).filter_by(
                        company_code=company_code).first()
                else:
                    existing = session.query(Member).filter_by(
                        org_name=org_name).first()
                exists = existing is not None
                if exists:
                    member_id = existing.id

            if exists and not overwrite:
                result["skipped"] += 1
                result["skipped_details"].append(
                    f"{row_idx}行目：{org_name}（既存レコードあり、上書きOFFのためスキップ）"
                )
                continue

            if exists:
                self._svc.update(member_id, data, "Excelインポートによる更新", staff_name)
                result["updated"] += 1
            else:
                self._svc.create(data, staff_name)
                result["added"] += 1

        return result


class ExportService:
    def __init__(self, engine):
        self._engine = engine

    def export_excel(self, members: list, output_path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "加入者名簿"
        ws.append(EXPORT_HEADERS)
        ins_order = ["ippan", "kensetsu_koyou", "ringyo", "kensetsu_genba", "kensetsu_jimusho"]
        for m in members:
            ins_map = {e.ins_type: e for e in m.insurance_entries}
            ins_cols = []
            for ins_type in ins_order:
                e = ins_map.get(ins_type)
                ins_cols += [
                    e.branch_number if e else "",
                    e.ins_number if e else "",
                    1 if (e and e.is_tokubetsu) else "",
                    1 if (e and e.is_ikkatsu) else "",
                ]
            addr_parts = (m.address or "").split("\n", 1)
            addr_mail_parts = (m.address_mail or "").split("\n", 1)
            ws.append([
                m.company_code or "",
                "○" if getattr(m, "is_member", True) else "",
                m.member_number or "",
                m.org_name, m.org_kana or "", m.dept_title or "",
                m.rep_name or "", m.rep_kana or "", m.email or "",
                m.tel_area or "", m.tel or "", m.fax_area or "", m.fax or "",
                m.postal_code or "",
                addr_parts[0],
                addr_parts[1] if len(addr_parts) > 1 else "",
                m.postal_code_mail or "",
                addr_mail_parts[0],
                addr_mail_parts[1] if len(addr_mail_parts) > 1 else "",
                m.mail_org_name or "",
                m.mail_dept_title or "",
                m.mail_person_name or "",
                m.employment_ins_no or "",
            ] + ins_cols + [
                m.note or "",
                m.registered_date.strftime("%Y-%m-%d") if m.registered_date else "",
            ] + [value for i in range(3) for value in (
                m.email_addresses[i].address if i < len(m.email_addresses) else "",
                m.email_addresses[i].label if i < len(m.email_addresses) else "",
            )])
        wb.save(output_path)

    @staticmethod
    def generate_template(output_path: str) -> None:
        """インポート用テンプレートExcelを生成する"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "加入者名簿"
        ws.append(EXPORT_HEADERS)
        wb.save(output_path)
