import pytest
from sqlalchemy import create_engine
import openpyxl
from app.database.models import Base, Member
from app.database.connection import get_session
from app.services.fee_service import FeeService
from app.services.fee_export_service import FeeExportService, HEADERS


@pytest.fixture
def engine():
    eng = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(eng)
    return eng


@pytest.fixture
def fee_svc(engine):
    return FeeService(engine)


@pytest.fixture
def export_svc(engine):
    return FeeExportService(engine)


def test_export_excel_writes_header_row(tmp_path, engine, fee_svc, export_svc):
    with get_session(engine) as session:
        session.add(Member(member_number="9001", org_name="A社", is_active=True, is_member=True))
    fee_svc.generate_records(2026)
    out = tmp_path / "out.xlsx"
    count = export_svc.export_excel(2026, str(out))
    assert count == 1
    wb = openpyxl.load_workbook(str(out))
    ws = wb["全件一覧"]
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == HEADERS
    assert len(HEADERS) == 26


def test_export_excel_writes_member_and_non_member_rows(tmp_path, engine, fee_svc, export_svc):
    with get_session(engine) as session:
        session.add(Member(member_number="9001", org_name="A社", is_active=True, is_member=True))
        session.add(Member(member_number="9002", org_name="B社", is_active=True, is_member=False))
    fee_svc.generate_records(2026)
    out = tmp_path / "out.xlsx"
    export_svc.export_excel(2026, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["全件一覧"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    kubun_by_org = {r[3]: r[4] for r in rows}
    assert kubun_by_org["A社"] == "会員"
    assert kubun_by_org["B社"] == "非会員"


def test_export_excel_diff_blank_when_unpaid(tmp_path, engine, fee_svc, export_svc):
    with get_session(engine) as session:
        session.add(Member(member_number="9001", org_name="A社", is_active=True, is_member=True))
    fee_svc.generate_records(2026)
    out = tmp_path / "out.xlsx"
    export_svc.export_excel(2026, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["全件一覧"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[23] in (None, "")  # 差額列（0始まりで24列目）


def test_export_excel_handles_none_note_without_crashing(tmp_path, engine, fee_svc, export_svc):
    with get_session(engine) as session:
        session.add(Member(member_number="9001", org_name="A社", is_active=True, is_member=True))
    fee_svc.generate_records(2026)
    out = tmp_path / "out.xlsx"
    # note, payment_method 等が未設定(None)の状態でも例外を出さずに出力できること
    count = export_svc.export_excel(2026, str(out))
    assert count == 1


def test_export_excel_premium_and_total_amount_values(tmp_path, engine, fee_svc, export_svc):
    with get_session(engine) as session:
        session.add(Member(member_number="9001", org_name="A社", is_active=True, is_member=True))
    fee_svc.generate_records(2026)
    with get_session(engine) as session:
        from app.database.models import AnnualFeeRecord
        record = session.query(AnnualFeeRecord).filter_by(fiscal_year=2026).first()
        record_id = record.id
    fee_svc.update(record_id, {"premium_branch_0": 200000})
    out = tmp_path / "out.xlsx"
    export_svc.export_excel(2026, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["全件一覧"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    # 概算保険料合計(index 10), 請求合計(index 16)
    assert row[10] == 200000
    assert row[16] == 11000
