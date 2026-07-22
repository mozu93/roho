from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine

from app.database.connection import get_session
from app.database.models import Base, Member
from app.services.bank_account_excel_service import BankAccountExcelService, HEADERS
from app.services.bank_account_service import BankAccountService


@pytest.fixture
def engine():
    value = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(value)
    with get_session(value) as session:
        session.add(Member(company_code=1001, org_name="サンプル商事"))
    return value


def _write(path: Path, rows: list[list]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "振込先口座"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_import_adds_account_and_restores_leading_zero(engine, tmp_path):
    path = tmp_path / "accounts.xlsx"
    _write(path, [[
        1001, "サンプル商事", "", "有効", 155, "百五銀行", 7, "本店",
        1, "普通", 123456, "カ）サンプル",
    ]])
    result = BankAccountExcelService(engine).import_excel(str(path))
    assert result == {"added": 1, "updated": 0, "skipped": 0, "errors": []}
    with get_session(engine) as session:
        member = session.query(Member).filter_by(company_code=1001).one()
        account = member.bank_accounts[0]
        assert account.bank_code == "0155"
        assert account.branch_code == "007"
        assert account.account_number == "0123456"
        assert account.recipient_name_kana == "ｶ)ｻﾝﾌﾟﾙ"


def test_export_and_reimport_updates_account(engine, tmp_path):
    with get_session(engine) as session:
        member_id = session.query(Member).filter_by(company_code=1001).one().id
    account = BankAccountService(engine).create(member_id, {
        "bank_code": "0155", "bank_name": "百五銀行",
        "branch_code": "307", "branch_name": "旭が丘支店",
        "account_type": "1", "account_number": "0123456",
        "recipient_name_kana": "ｶ)ｻﾝﾌﾟﾙ", "is_enabled": True,
    })
    path = tmp_path / "export.xlsx"
    count = BankAccountExcelService(engine).export_excel(str(path))
    assert count == 1
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["振込先口座"]
    sheet.cell(2, HEADERS.index("支店名") + 1).value = "鈴鹿支店"
    workbook.save(path)
    result = BankAccountExcelService(engine).import_excel(str(path))
    assert result["updated"] == 1
    assert BankAccountService(engine).get(account.id).branch_name == "鈴鹿支店"


def test_import_reports_unknown_customer_without_stopping(engine, tmp_path):
    path = tmp_path / "error.xlsx"
    _write(path, [[
        9999, "不明", "", "有効", "0155", "百五銀行", "307", "旭が丘支店",
        "1", "普通", "0123456", "ｶ)ｻﾝﾌﾟﾙ",
    ]])
    result = BankAccountExcelService(engine).import_excel(str(path))
    assert result["skipped"] == 1
    assert "顧客コード9999" in result["errors"][0]
